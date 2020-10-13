import pprint
import os, time, math, threading

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl as exl

from libs import converter as cvt
from libs import parallel

DRIVER_PATH = os.environ['DRIVER_PATH']
GOOGLE_URL = 'https://google.com/'



# csvから取得〜sheetに書き込み
def write_to_sheet(book, sheet, row, company) -> None:
    """シート書き込み"""
    try:
        sheet.cell(row=row, column=1, value=company['id'])
        sheet.cell(row=row, column=2, value=company['name'])
        sheet.cell(row=row, column=3, value=company['url'])
        sheet.cell(row=row, column=4, value=company['charge'])
        sheet.cell(row=row, column=5, value=company['appointment'])
        sheet.cell(row=row, column=6, value=company['phone_number'])
        sheet.cell(row=row, column=7, value=company['phone_number_2'])
        return
    except Exception:
        return

def get_phone_number(company_name) -> str:
    """電話番号の更新"""

    def get_driver():
        """Chromeのdriverを取得"""
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--lang=ja') # configure language
        driver = webdriver.Chrome(DRIVER_PATH, options=options)
        return driver

    driver = get_driver()
    serach_word = f'{company_name} 電話番号'
    # start
    driver.get(GOOGLE_URL)
    serach_form = driver.find_element_by_name('q')
    serach_form.send_keys(serach_word)
    serach_form.submit()
    try:
        phone_number = driver.find_element_by_class_name('mw31Ze').text
    except Exception:
        phone_number = ''
    # end
    finally:
        driver.quit()
        return phone_number


def update_company_list(company_list) -> list:
    """会社情報の取得・更新"""

    def add_phone_number(company, company_index) -> None:
        """会社情報の更新と追加"""
        company['id'] = company_index
        company['phone_number_2'] = get_phone_number(company['name'])
        result.append(company)
        print(company_index)
        return

    # [並列処理]
    result = []
    i = 0
    thread_range = 10
    loop_times = math.ceil(len(company_list)/thread_range)
    # 総レコード数をthread_rangeで割った回数分ループ
    for _ in range(loop_times):
        threads = []
        for j in range(thread_range):
            company_index = i*thread_range + j
            company = company_list[company_index]
            t = threading.Thread(
                target=add_phone_number,
                args=(company, company_index))
            t.setDaemon(True)
            t.start()
            threads.append(t)
        parallel.wait_all_threads(threads)
        i += 1

    # idで並び順を整理
    return sorted(result, key=lambda x: x['id'])


def main() -> None:

    # 1. ブック取得
    book_path: str = './_storage/output.xlsx'
    book = exl.load_workbook(book_path)
    sheet = book.worksheets[0]
    input_file_path: str = './_storage/company_list.csv'
    dict_keys: list = ['name', 'url', 'charge', 'appointment', 'phone_number']
    company_list: list = cvt.csv_to_dicts(
        csv_path=input_file_path,
        dict_keys=dict_keys
    )
    company_list: list = company_list[27000:27500]

    # 2. company_listを更新
    updated_company_list = update_company_list(company_list)

    # 3. シートにラベルの追加
    sheet_label: dict = {
        'id': 'No.',
        'name': '会社名',
        'url': 'URL',
        'charge': '担当者',
        'appointment': '結果',
        'phone_number': '電話番号',
        'phone_number_2': '取得電話番号'
    }
    updated_company_list.insert(0, sheet_label)

    # 4. シートに書込み
    for i, company in enumerate(updated_company_list):
        write_to_sheet(book, sheet, row=i+1, company=company)
        book.save(book_path)



if __name__ == '__main__':
    # エクセルをCSVに変換
    cvt.excel_to_csv(
        xlsx_path='_storage/kaden.xlsx',
        csv_path='./_storage/kaden.csv',
        sheet_name='！新規獲得',
        usecols=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11],
        index_col=1)
    # エンコーディング変換
    # print(cvt.get_encoding('./_storage/sample.csv'))
    # cvt.convert_encoding(
    #     './_storage/sample.csv',
    #     './_storage/result.csv',
    #     from_encoding='CP932',
    #     to_encoding='utf-8'
    # )


    # CSVデータからエクセルを作成
    # cvt.create_excel_book('./_storage/output.xlsx')
    # main()
